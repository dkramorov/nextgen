# -*- coding: utf-8 -*-
import os
import logging
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook

from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render

logger = logging.getLogger(__name__)

def index(request):
    context = {
      'name':'test_xlsx',
    }
    return render(request, 'excelka/test_xlsx.html', context)
    #return HttpResponse("Hi")

#################################################
# Загрузка файла xlsx
# Выгрузка в xlsx
#################################################
def upload_xlsx(request):
    context = {
      'name':'upload_xlsx',
    }
    f = request.FILES.get("file")
    context['name'] = f.name

    # ---------------------------
    # file['names'], file['rows']
    # ---------------------------
    context['file'] = read_workbook(f)
    output = create_workbook(
        context['file']['names'],
        context['file']['rows']
    )

    name, ext = f.name.split(".", 1)
    fname = "{}.xlsx".format(name)
    dest = os.path.join(settings.MEDIA_ROOT, fname)
    with open(dest,'wb+') as out:
        out.write(output.read())
    context['dest'] = "{}{}".format(settings.MEDIA_URL, fname)

    #response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #response['Content-Disposition'] = "attachment; filename=sales_report.xlsx"
    #return response

    return JsonResponse(context)

#################################################
# openpyxl - xlsx
# xlswriter - xlsx, pip install XlsxWriter
# xlrd - xls
# xlwt - xls
#################################################

#################################################
# Чтение эксельки xlsx
#################################################
def read_workbook(excel_file):
    result = {}
    result['errors'] = []
    # --------------------------------------------------
    # data_only=True - чтобы ибучие формулы не писались,
    # а то вывод будет пзд кривой =M2&" "&N2&" всякие
    # --------------------------------------------------
    wb = load_workbook(
        BytesIO(excel_file.read()),
        data_only=True,
    )
    sheet = wb.active

    rows = list(sheet.rows)
    if not len(rows) > 0:
        result['errors'].append('Файл пустой')
        return result
    # ------------------
    # Заголовки столбцов
    # Пустые ПНХ
    # ------------------
    names = [cell.value for cell in rows[0] if cell.value]
    # -----------------------------------
    # Если в имена добавить валидацию,
    # то удобно будет приводить
    # к нужному виду, сообщать об ошибках
    # -----------------------------------
    result['names'] = names
    result['rows'] = []
    result['rows_length'] = {name:0 for name in names}
    # -------------------------
    # Проходим по каждой строке
    # Кроме первой
    # -------------------------
    rows = sheet.rows
    next(rows) # Генератор (пропускаем первую строку)
    for row in rows:
        # ----------------------------
        # Каждую строку пишем по names
        # ----------------------------
        line = []
        for name in names:
          # ------------------------------
          # По индексу определяем значение
          # ------------------------------
          ind = names.index(name)
          value = row[ind].value or ""
          len_value = len("{}".format(value))
          if result['rows_length'][name] < len_value:
              result['rows_length'][name] = len_value
          line.append(value)

        # ---------------------------
        # Добавляем линию в результат
        # ---------------------------
        result['rows'].append(line)
    return result

#################################################
# Создание эксельки xlsx
# names - заголовки таблицы, rows - данные
# sheet_name - имя листа
#################################################
def create_workbook(names, rows, sheet_name=None):
    # -------------------------
    # Будем создавать в памяти
    # excel файл для скачивания
    # -------------------------
    output = BytesIO()
    book = xlsxwriter.Workbook(
        output,
        # ---------------------------
        # Не использовать гиперссылки
        # ---------------------------
        {'strings_to_urls': False},
    )

    if not sheet_name:
        sheet_name = 'Лист 1'
    sheet = book.add_worksheet(sheet_name)
    # -----------------------------------
    # row_number - строка по вертикали
    # cell_number - ячейка по горизонтали
    # -----------------------------------
    row_number = 0
    cell_number = 0

    # -----------------
    # Заголовок таблицы
    # -----------------
    for name in enumerate(names):
        sheet.write(row_number, name[0], name[1])

    for row in rows:
        row_number += 1
        # ------------------
        # Может обыграть его
        # следующим листом?
        # ------------------
        if row_number > 65500:
            msg = 'К сожалению, файл был обрезан из-за большого количества позиций'
            sheet.merge_range('A{}:H{}'.format(row_number, row_number), msg)
            break

        for cell in enumerate(row):
            value = cell[1] or ''
            sheet.write(row_number, cell[0], value)

    book.close()
    output.seek(0)

    return output

